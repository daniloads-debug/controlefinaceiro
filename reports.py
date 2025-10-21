import pandas as pd
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from datetime import datetime, timedelta
import plotly.graph_objects as go
import plotly.express as px
import io
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import PieChart, Reference, BarChart

class ReportGenerator:
    def __init__(self, database, analytics):
        self.db = database
        self.analytics = analytics
        self.styles = getSampleStyleSheet()
    
    def generate_monthly_excel_report(self, year, month, filename=None):
        """Gera relatório mensal em Excel"""
        if not filename:
            filename = f"relatorio_mensal_{year}_{month:02d}.xlsx"
        
        # Criar workbook
        wb = Workbook()
        
        # Aba 1: Resumo
        ws_summary = wb.active
        ws_summary.title = "Resumo"
        
        # Cabeçalho
        ws_summary['A1'] = f"Relatório Financeiro - {month:02d}/{year}"
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws_summary['A1'].font = Font(color="FFFFFF", size=16, bold=True)
        
        # Obter dados
        insights = self.analytics.calculate_category_insights(year, month)
        monthly_data = self.db.get_monthly_summary(year, month)
        
        if insights:
            # Métricas principais
            ws_summary['A3'] = "Métricas Principais"
            ws_summary['A3'].font = Font(bold=True)
            
            metrics = [
                ["Total de Receitas", f"R$ {insights['total_income']:,.2f}"],
                ["Total de Despesas", f"R$ {insights['total_expenses']:,.2f}"],
                ["Saldo do Mês", f"R$ {insights['balance']:,.2f}"],
                ["Taxa de Poupança", f"{insights['savings_rate']:.1f}%"]
            ]
            
            for i, (label, value) in enumerate(metrics, start=4):
                ws_summary[f'A{i}'] = label
                ws_summary[f'B{i}'] = value
                ws_summary[f'A{i}'].font = Font(bold=True)
        
        # Aba 2: Transações
        ws_transactions = wb.create_sheet("Transações")
        
        # Obter transações do mês
        start_date = datetime(year, month, 1).date()
        if month == 12:
            end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
        
        transactions = self.db.get_transactions(start_date, end_date)
        
        if not transactions.empty:
            # Cabeçalhos
            headers = ["Data", "Descrição", "Categoria", "Valor", "Tipo"]
            for i, header in enumerate(headers, start=1):
                cell = ws_transactions.cell(row=1, column=i, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            # Dados
            for row_idx, (_, transaction) in enumerate(transactions.iterrows(), start=2):
                ws_transactions.cell(row=row_idx, column=1, value=transaction['date'])
                ws_transactions.cell(row=row_idx, column=2, value=transaction['description'])
                ws_transactions.cell(row=row_idx, column=3, value=transaction['category'])
                ws_transactions.cell(row=row_idx, column=4, value=f"R$ {transaction['amount']:,.2f}")
                ws_transactions.cell(row=row_idx, column=5, value=transaction['type'])
        
        # Aba 3: Análise por Categoria
        ws_categories = wb.create_sheet("Por Categoria")
        
        if not monthly_data.empty:
            # Cabeçalhos
            headers = ["Categoria", "Tipo", "Total", "Quantidade"]
            for i, header in enumerate(headers, start=1):
                cell = ws_categories.cell(row=1, column=i, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            # Dados
            for row_idx, (_, category_data) in enumerate(monthly_data.iterrows(), start=2):
                ws_categories.cell(row=row_idx, column=1, value=category_data['category'])
                ws_categories.cell(row=row_idx, column=2, value=category_data['type'])
                ws_categories.cell(row=row_idx, column=3, value=f"R$ {category_data['total']:,.2f}")
                ws_categories.cell(row=row_idx, column=4, value=category_data['count'])
        
        # Salvar arquivo
        wb.save(filename)
        return filename
    
    def generate_pdf_report(self, year, month, filename=None):
        """Gera relatório mensal em PDF"""
        if not filename:
            filename = f"relatorio_mensal_{year}_{month:02d}.pdf"
        
        doc = SimpleDocTemplate(filename, pagesize=A4)
        story = []
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Centralizado
        )
        
        title = Paragraph(f"Relatório Financeiro - {month:02d}/{year}", title_style)
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Obter dados
        insights = self.analytics.calculate_category_insights(year, month)
        
        if insights:
            # Resumo financeiro
            summary_style = ParagraphStyle(
                'Summary',
                parent=self.styles['Normal'],
                fontSize=12,
                spaceAfter=10
            )
            
            story.append(Paragraph("Resumo Financeiro", self.styles['Heading2']))
            
            summary_data = [
                ['Métrica', 'Valor'],
                ['Total de Receitas', f"R$ {insights['total_income']:,.2f}"],
                ['Total de Despesas', f"R$ {insights['total_expenses']:,.2f}"],
                ['Saldo do Mês', f"R$ {insights['balance']:,.2f}"],
                ['Taxa de Poupança', f"{insights['savings_rate']:.1f}%"]
            ]
            
            summary_table = Table(summary_data)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(summary_table)
            story.append(Spacer(1, 20))
            
            # Top categorias de despesa
            if insights['top_expense_categories']:
                story.append(Paragraph("Top 5 Categorias de Despesa", self.styles['Heading2']))
                
                category_data = [['Categoria', 'Valor', 'Quantidade']]
                for cat in insights['top_expense_categories'][:5]:
                    category_data.append([
                        cat['category'],
                        f"R$ {cat['total']:,.2f}",
                        str(cat['count'])
                    ])
                
                category_table = Table(category_data)
                category_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(category_table)
                story.append(Spacer(1, 20))
        
        # Transações recentes
        start_date = datetime(year, month, 1).date()
        if month == 12:
            end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
        
        transactions = self.db.get_transactions(start_date, end_date)
        
        if not transactions.empty:
            story.append(Paragraph("Últimas Transações", self.styles['Heading2']))
            
            # Mostrar apenas as 10 mais recentes
            recent_transactions = transactions.head(10)
            
            transaction_data = [['Data', 'Descrição', 'Categoria', 'Valor', 'Tipo']]
            for _, transaction in recent_transactions.iterrows():
                transaction_data.append([
                    transaction['date'],
                    transaction['description'][:30] + '...' if len(transaction['description']) > 30 else transaction['description'],
                    transaction['category'],
                    f"R$ {transaction['amount']:,.2f}",
                    transaction['type']
                ])
            
            transaction_table = Table(transaction_data)
            transaction_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(transaction_table)
        
        # Rodapé
        story.append(Spacer(1, 30))
        footer_style = ParagraphStyle(
            'Footer',
            parent=self.styles['Normal'],
            fontSize=10,
            alignment=1,
            textColor=colors.grey
        )
        
        footer = Paragraph(f"Relatório gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}", footer_style)
        story.append(footer)
        
        # Construir PDF
        doc.build(story)
        return filename
    
    def generate_annual_projection_report(self, filename=None):
        """Gera relatório de projeção anual"""
        if not filename:
            filename = f"projecao_anual_{datetime.now().year}.pdf"
        
        doc = SimpleDocTemplate(filename, pagesize=A4)
        story = []
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1
        )
        
        title = Paragraph(f"Projeção Anual - {datetime.now().year}", title_style)
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Obter projeções
        projections = self.analytics.predict_annual_projection()
        
        if projections:
            story.append(Paragraph("Projeções por Categoria", self.styles['Heading2']))
            
            projection_data = [['Categoria', 'Projeção Anual', 'Média Mensal', 'Tendência', 'Confiança']]
            total_projected = 0
            
            for category, data in projections.items():
                projection_data.append([
                    category,
                    f"R$ {data['annual_total']:,.2f}",
                    f"R$ {data['average_monthly']:,.2f}",
                    data['trend'],
                    f"{data['confidence']*100:.0f}%"
                ])
                total_projected += data['annual_total']
            
            # Adicionar total
            projection_data.append([
                'TOTAL',
                f"R$ {total_projected:,.2f}",
                f"R$ {total_projected/12:,.2f}",
                '-',
                '-'
            ])
            
            projection_table = Table(projection_data)
            projection_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(projection_table)
            story.append(Spacer(1, 20))
            
            # Disclaimer
            disclaimer_style = ParagraphStyle(
                'Disclaimer',
                parent=self.styles['Normal'],
                fontSize=10,
                textColor=colors.grey,
                leftIndent=20,
                rightIndent=20
            )
            
            disclaimer = Paragraph(
                "<b>Aviso:</b> As projeções são baseadas em dados históricos e algoritmos de machine learning. "
                "Os valores reais podem variar conforme mudanças nos hábitos de consumo, situação econômica "
                "e outros fatores externos. Use estas informações como referência para planejamento financeiro.",
                disclaimer_style
            )
            story.append(disclaimer)
        
        else:
            story.append(Paragraph("Dados insuficientes para gerar projeções.", self.styles['Normal']))
        
        # Construir PDF
        doc.build(story)
        return filename
    
    def export_transactions_to_excel(self, start_date=None, end_date=None, filename=None):
        """Exporta todas as transações para Excel"""
        if not filename:
            filename = f"transacoes_export_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        # Obter transações
        transactions = self.db.get_transactions(start_date, end_date)
        
        if transactions.empty:
            return None
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Transações"
        
        # Cabeçalhos
        headers = ["ID", "Data", "Descrição", "Valor", "Categoria", "Tipo", "Criado em"]
        for i, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=i, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Dados
        for row_idx, (_, transaction) in enumerate(transactions.iterrows(), start=2):
            ws.cell(row=row_idx, column=1, value=transaction['id'])
            ws.cell(row=row_idx, column=2, value=transaction['date'])
            ws.cell(row=row_idx, column=3, value=transaction['description'])
            ws.cell(row=row_idx, column=4, value=transaction['amount'])
            ws.cell(row=row_idx, column=5, value=transaction['category'])
            ws.cell(row=row_idx, column=6, value=transaction['type'])
            ws.cell(row=row_idx, column=7, value=transaction.get('created_at', ''))
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar arquivo
        wb.save(filename)
        return filename